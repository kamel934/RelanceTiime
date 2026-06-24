from __future__ import annotations

import csv
from pathlib import Path

from openpyxl import load_workbook

from outil_factures_manquantes import (
    APP_VERSION,
    CUSTOMER_RECEIPTS_WITHOUT_SALES_INVOICE,
    CUSTOMER_SALES_INVOICES_WITHOUT_PAYMENT,
    INVOICES_WITHOUT_PAYMENT,
    LEDGER_CUSTOMER,
    LEDGER_SUPPLIER,
    OutputFormats,
    PAYMENTS_WITHOUT_INVOICE,
    derive_client_year,
    detect_ledger_type,
    process_file,
)


HEADERS = [
    "N° Compte",
    "N° Compte Auxiliaire",
    "Libellé du compte",
    "Date",
    "Journal",
    "N° de pièce",
    "Libellé mouvement",
    "Lettrage",
    "Montant Débit",
    "Montant Crédit",
    "Solde ",
]


def write_csv(path: Path, rows: list[dict[str, str]]) -> Path:
    with path.open("w", encoding="cp1252", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=HEADERS, delimiter=";")
        writer.writeheader()
        for values in rows:
            row = {header: "" for header in HEADERS}
            row.update(values)
            writer.writerow(row)
    return path


def supplier_csv(tmp_path: Path) -> Path:
    return write_csv(
        tmp_path / "grand_livre_2025-01-01_2025-12-31_fournisseur_test.csv",
        [
            {
                "N° Compte": "40100000",
                "Libellé du compte": "FOURNISSEUR TEST",
                "Date": "10/01/2025",
                "Journal": "BQ",
                "N° de pièce": "",
                "Libellé mouvement": "PAIEMENT FOURNISSEUR",
                "Montant Débit": "100,00",
                "Montant Crédit": "0,00",
            },
            {
                "N° Compte": "40100000",
                "Libellé du compte": "FOURNISSEUR TEST",
                "Date": "11/01/2025",
                "Journal": "AO",
                "N° de pièce": "",
                "Libellé mouvement": "MOUVEMENT A EXCLURE",
                "Montant Débit": "50,00",
                "Montant Crédit": "0,00",
            },
            {
                "N° Compte": "40100000",
                "Libellé du compte": "FOURNISSEUR TEST",
                "Date": "12/01/2025",
                "Journal": "AC",
                "N° de pièce": "FAC001",
                "Libellé mouvement": "FACTURE FOURNISSEUR",
                "Montant Débit": "0,00",
                "Montant Crédit": "250,00",
            },
        ],
    )


def customer_csv(tmp_path: Path) -> Path:
    return write_csv(
        tmp_path / "grand_livre_2025-01-01_2025-12-31_client_test.csv",
        [
            {
                "N° Compte": "41100000",
                "Libellé du compte": "CLIENT TEST",
                "Date": "10/01/2025",
                "Journal": "BQ",
                "N° de pièce": "",
                "Libellé mouvement": "ENCAISSEMENT CLIENT",
                "Montant Débit": "0,00",
                "Montant Crédit": "120,00",
            },
            {
                "N° Compte": "41100000",
                "Libellé du compte": "CLIENT TEST",
                "Date": "11/01/2025",
                "Journal": "VI",
                "N° de pièce": "V001",
                "Libellé mouvement": "FACTURE DE VENTE",
                "Montant Débit": "300,00",
                "Montant Crédit": "0,00",
            },
            {
                "N° Compte": "41100000",
                "Libellé du compte": "CLIENT TEST",
                "Date": "12/01/2025",
                "Journal": "VT",
                "N° de pièce": "AV001",
                "Libellé mouvement": "AVOIR DE VENTE",
                "Montant Débit": "0,00",
                "Montant Crédit": "40,00",
            },
        ],
    )


def test_version_and_filename_with_windows_suffix(tmp_path: Path) -> None:
    assert APP_VERSION == "1.0.0"
    path = tmp_path / "grand_livre_2024-12-01_2025-12-31_societe_test (1).csv"
    assert derive_client_year(path) == ("SOCIETE TEST", "2024-2025")


def test_supplier_detection_and_exports(tmp_path: Path) -> None:
    source = supplier_csv(tmp_path)
    assert detect_ledger_type(source) == LEDGER_SUPPLIER

    payments = process_file(source, PAYMENTS_WITHOUT_INVOICE, OutputFormats(True, True))
    assert payments.row_count == 1
    assert payments.total_amount == 100
    assert payments.pdf_output.read_bytes().startswith(b"%PDF-")
    payment_sheet = load_workbook(payments.excel_output).active
    assert [cell.value for cell in payment_sheet[1]] == [
        "Libellé du compte",
        "Date",
        "Journal",
        "Libellé mouvement",
        "Paiements",
        "Remarque",
    ]

    invoices = process_file(source, INVOICES_WITHOUT_PAYMENT, OutputFormats(True, False))
    assert invoices.row_count == 1
    assert invoices.total_amount == 250


def test_customer_detection_and_exports(tmp_path: Path) -> None:
    source = customer_csv(tmp_path)
    assert detect_ledger_type(source) == LEDGER_CUSTOMER

    receipts = process_file(
        source,
        CUSTOMER_RECEIPTS_WITHOUT_SALES_INVOICE,
        OutputFormats(True, True),
    )
    assert receipts.row_count == 1
    assert receipts.total_amount == 120
    assert receipts.pdf_output.read_bytes().startswith(b"%PDF-")
    receipt_sheet = load_workbook(receipts.excel_output).active
    assert [cell.value for cell in receipt_sheet[1]] == [
        "Libellé du compte",
        "Date",
        "Journal",
        "Libellé mouvement",
        "Encaissements",
        "Remarque",
    ]

    invoices = process_file(
        source,
        CUSTOMER_SALES_INVOICES_WITHOUT_PAYMENT,
        OutputFormats(True, False),
    )
    assert invoices.row_count == 2
    assert invoices.total_amount == 300
    invoice_sheet = load_workbook(invoices.excel_output).active
    assert [cell.value for cell in invoice_sheet[1]] == [
        "Libellé du compte",
        "Date",
        "Journal",
        "Libellé mouvement",
        "Factures",
        "Avoirs",
        "Remarque",
    ]
    assert invoice_sheet.cell(3, 7).value == "Avoir de vente sans rmbrsmt"
