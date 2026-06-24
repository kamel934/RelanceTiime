from __future__ import annotations

import csv
import json
import os
import re
import sys
import tempfile
import unicodedata
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from tkinter import Tk, filedialog, messagebox, simpledialog
import tkinter as tk

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.table import Table
from openpyxl.utils import column_index_from_string, get_column_letter

try:
    from tkinterdnd2 import DND_FILES, TkinterDnD
except Exception:
    DND_FILES = None
    TkinterDnD = None


EXCLUDED_JOURNALS = {"AT", "AC", "AO", "OD", "CA", "AN", "RV"}
APP_VERSION = "1.0.0"
SALE_JOURNALS = {"VI", "VE", "VT"}
CUSTOMER_RECEIPT_EXCLUDED_JOURNALS = EXCLUDED_JOURNALS | SALE_JOURNALS
SUPPLIER_DETECTION_JOURNALS = {"AC", "AT", "AO"}
LEDGER_AUTO = "auto"
LEDGER_SUPPLIER = "supplier"
LEDGER_CUSTOMER = "customer"
LEDGER_MODES = {LEDGER_AUTO, LEDGER_SUPPLIER, LEDGER_CUSTOMER}
ROLE_FIRST = "first"
ROLE_SECOND = "second"
MOVEMENT_CHARS_PER_LINE = 64
BASE_ROW_HEIGHT = 22
WRAPPED_LINE_HEIGHT = 18
ROW_HEIGHT_PADDING = 4
MAX_ROW_HEIGHT = 84
REQUIRED_COLUMNS = {
    "libelle_du_compte": "Libellé du compte",
    "date": "Date",
    "journal": "Journal",
    "n_de_piece": "N° de pièce",
    "libelle_mouvement": "Libellé mouvement",
    "montant_debit": "Montant Débit",
    "montant_credit": "Montant Crédit",
}


@dataclass
class OutputInfo:
    source: Path
    excel_output: Path | None
    pdf_output: Path | None
    row_count: int
    amount_label: str
    total_amount: float
    treatment_label: str


@dataclass(frozen=True)
class OutputFormats:
    excel: bool
    pdf: bool

    @property
    def has_any(self) -> bool:
        return self.excel or self.pdf


@dataclass(frozen=True)
class Treatment:
    key: str
    ledger_type: str
    role: str
    label: str
    output_label: str
    sheet_title: str
    table_name: str
    amount_label: str
    primary_column: str
    secondary_column: str


PAYMENTS_WITHOUT_INVOICE = Treatment(
    key="supplier_payments",
    ledger_type=LEDGER_SUPPLIER,
    role=ROLE_FIRST,
    label="Paiements sans facture",
    output_label="PAIEMENTS SANS FACTURE",
    sheet_title="Paiements sans facture",
    table_name="PaiementsSansFacture",
    amount_label="total paiements",
    primary_column="Paiements",
    secondary_column="Rmbrsmts",
)
INVOICES_WITHOUT_PAYMENT = Treatment(
    key="supplier_invoices",
    ledger_type=LEDGER_SUPPLIER,
    role=ROLE_SECOND,
    label="Factures sans paiements",
    output_label="FACTURES SANS PAIEMENTS",
    sheet_title="Factures sans paiements",
    table_name="FacturesSansPaiements",
    amount_label="total factures",
    primary_column="Factures",
    secondary_column="Avoirs",
)
CUSTOMER_RECEIPTS_WITHOUT_SALES_INVOICE = Treatment(
    key="customer_receipts",
    ledger_type=LEDGER_CUSTOMER,
    role=ROLE_FIRST,
    label="Encaissements sans facture de ventes",
    output_label="ENCAISSEMENTS SANS FACTURE DE VENTES",
    sheet_title="Encaissements sans facture",
    table_name="EncaissementsSansFacture",
    amount_label="total encaissements",
    primary_column="Encaissements",
    secondary_column="Remboursements",
)
CUSTOMER_SALES_INVOICES_WITHOUT_PAYMENT = Treatment(
    key="customer_invoices",
    ledger_type=LEDGER_CUSTOMER,
    role=ROLE_SECOND,
    label="Factures de ventes sans paiement",
    output_label="FACTURES DE VENTES SANS PAIEMENT",
    sheet_title="Factures ventes sans paiement",
    table_name="FacturesVentesSansPaiement",
    amount_label="total factures",
    primary_column="Factures",
    secondary_column="Avoirs",
)
TREATMENTS_BY_LEDGER_ROLE = {
    (LEDGER_SUPPLIER, ROLE_FIRST): PAYMENTS_WITHOUT_INVOICE,
    (LEDGER_SUPPLIER, ROLE_SECOND): INVOICES_WITHOUT_PAYMENT,
    (LEDGER_CUSTOMER, ROLE_FIRST): CUSTOMER_RECEIPTS_WITHOUT_SALES_INVOICE,
    (LEDGER_CUSTOMER, ROLE_SECOND): CUSTOMER_SALES_INVOICES_WITHOUT_PAYMENT,
}
ALL_TREATMENTS = (PAYMENTS_WITHOUT_INVOICE, INVOICES_WITHOUT_PAYMENT)
EXCEL_ONLY = OutputFormats(excel=True, pdf=False)
EXCEL_AND_PDF = OutputFormats(excel=True, pdf=True)


def settings_path() -> Path:
    base = Path(os.environ.get("APPDATA") or Path.home() / "AppData" / "Roaming")
    return base / "RelanceTiime" / "settings.json"


def read_settings() -> dict[str, object]:
    path = settings_path()
    if not path.exists():
        return {}
    try:
        data = json.loads(path.read_text(encoding="utf-8-sig"))
    except Exception:
        return {}
    return data if isinstance(data, dict) else {}


def write_settings(data: dict[str, object]) -> None:
    path = settings_path()
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(data, indent=2), encoding="utf-8")


def load_saved_formats(default_formats: OutputFormats, invalid_fallback: OutputFormats = EXCEL_ONLY) -> OutputFormats:
    data = read_settings()
    if not data:
        return default_formats
    try:
        formats = OutputFormats(excel=bool(data.get("excel")), pdf=bool(data.get("pdf")))
    except Exception:
        return default_formats
    if not formats.has_any:
        return invalid_fallback
    return formats


def save_formats(formats: OutputFormats) -> None:
    data = read_settings()
    data["excel"] = formats.excel
    data["pdf"] = formats.pdf
    write_settings(data)


def load_saved_ledger_mode(default_mode: str = LEDGER_AUTO) -> str:
    mode = str(read_settings().get("ledger_mode") or default_mode)
    return mode if mode in LEDGER_MODES else default_mode


def save_ledger_mode(mode: str) -> None:
    if mode not in LEDGER_MODES:
        mode = LEDGER_AUTO
    data = read_settings()
    data["ledger_mode"] = mode
    write_settings(data)


def resource_path(name: str) -> Path:
    base = Path(getattr(sys, "_MEIPASS", Path(__file__).resolve().parent))
    return base / name


def norm(value: str) -> str:
    text = "".join(
        ch for ch in unicodedata.normalize("NFKD", str(value or ""))
        if not unicodedata.combining(ch)
    )
    return re.sub(r"[^a-z0-9]+", "_", text.lower()).strip("_")


def parse_money(value: str | None) -> float:
    text = str(value or "").replace("\xa0", " ").strip()
    if not text:
        return 0.0
    text = text.replace(" ", "").replace(",", ".")
    return float(text)


def format_amount(value: float) -> str:
    return f"{value:,.2f}".replace(",", " ").replace(".", ",")


def write_error_log(message: str) -> None:
    try:
        log_path = Path(tempfile.gettempdir()) / "relance_tiime_error.log"
        log_path.write_text(message, encoding="utf-8")
    except Exception:
        pass


def parse_date(value: str | None) -> datetime | None:
    text = str(value or "").strip()
    if not text:
        return None
    for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y"):
        try:
            return datetime.strptime(text, fmt)
        except ValueError:
            pass
    raise ValueError(f"Date invalide : {text}")


def derive_client_year(csv_path: Path, root: Tk | None = None) -> tuple[str, str]:
    match = re.match(
        r"grand_livre_(\d{4})-\d{2}-\d{2}_(\d{4})-\d{2}-\d{2}_(.+?)(?: \(\d+\))?\.csv$",
        csv_path.name,
        flags=re.IGNORECASE,
    )
    if match:
        start_year, end_year, client_slug = match.groups()
        year = start_year if start_year == end_year else f"{start_year}-{end_year}"
        client = client_slug.replace("_", " ").replace("-", " ").upper().strip()
        return client, year

    if root is None:
        raise ValueError(
            "Nom de fichier non standard. Format attendu : "
            "grand_livre_2025-01-01_2025-12-31_nom_client.csv"
        )

    client = simpledialog.askstring("Client", "Nom du client :", parent=root)
    year = simpledialog.askstring("Année", "Année ou période :", parent=root)
    if not client or not year:
        raise ValueError("Client et année sont obligatoires pour ce fichier.")
    return client.upper().strip(), year.strip()


def unique_path(path: Path) -> Path:
    if not path.exists():
        return path
    stem, suffix = path.stem, path.suffix
    index = 2
    while True:
        candidate = path.with_name(f"{stem} - {index}{suffix}")
        if not candidate.exists():
            return candidate
        index += 1


def unique_output_paths(directory: Path, stem: str, formats: OutputFormats) -> tuple[Path | None, Path | None]:
    index = 1
    while True:
        suffix = "" if index == 1 else f" - {index}"
        excel_path = directory / f"{stem}{suffix}.xlsx" if formats.excel else None
        pdf_path = directory / f"{stem}{suffix}.pdf" if formats.pdf else None
        candidates = [path for path in (excel_path, pdf_path) if path is not None]
        if all(not path.exists() for path in candidates):
            return excel_path, pdf_path
        index += 1


def get_header_map(reader: csv.DictReader) -> dict[str, str]:
    if not reader.fieldnames:
        raise ValueError("Le CSV ne contient pas d'en-têtes.")
    header_map = {norm(header): header for header in reader.fieldnames}
    missing = [label for key, label in REQUIRED_COLUMNS.items() if key not in header_map]
    if missing:
        raise ValueError("Colonnes manquantes : " + ", ".join(missing))
    return header_map


def detect_ledger_type(csv_path: Path) -> str:
    with csv_path.open("r", encoding="cp1252", newline="") as handle:
        reader = csv.DictReader(handle, delimiter=";")
        header_map = get_header_map(reader)
        account_header = {norm(header): header for header in reader.fieldnames or []}.get("n_compte")
        journal_header = header_map["journal"]

        supplier_accounts = 0
        customer_accounts = 0
        supplier_journals = 0
        customer_journals = 0

        for raw in reader:
            if account_header:
                account = re.sub(r"\D+", "", str(raw.get(account_header, "") or ""))
                if account.startswith("401"):
                    supplier_accounts += 1
                elif account.startswith("411"):
                    customer_accounts += 1

            journal = str(raw.get(journal_header, "") or "").strip().upper()
            if journal in SUPPLIER_DETECTION_JOURNALS:
                supplier_journals += 1
            elif journal in SALE_JOURNALS:
                customer_journals += 1

    if supplier_accounts > customer_accounts:
        return LEDGER_SUPPLIER
    if customer_accounts > supplier_accounts:
        return LEDGER_CUSTOMER
    if supplier_journals > customer_journals:
        return LEDGER_SUPPLIER
    if customer_journals > supplier_journals:
        return LEDGER_CUSTOMER

    raise ValueError(
        "Impossible de reconnaître automatiquement le grand livre. "
        "Choisissez Fournisseur ou Client dans l'application."
    )


def resolve_ledger_type(csv_path: Path, ledger_mode: str = LEDGER_AUTO) -> str:
    if ledger_mode in {LEDGER_SUPPLIER, LEDGER_CUSTOMER}:
        return ledger_mode
    return detect_ledger_type(csv_path)


def resolve_treatments(csv_path: Path, roles: tuple[str, ...], ledger_mode: str = LEDGER_AUTO) -> tuple[Treatment, ...]:
    ledger_type = resolve_ledger_type(csv_path, ledger_mode)
    return tuple(TREATMENTS_BY_LEDGER_ROLE[(ledger_type, role)] for role in roles)


def estimate_wrapped_lines(value: object, chars_per_line: int = MOVEMENT_CHARS_PER_LINE) -> int:
    text = str(value or "").strip()
    if not text:
        return 1

    total_lines = 0
    for raw_segment in text.splitlines() or [""]:
        words = raw_segment.split()
        if not words:
            total_lines += 1
            continue

        line_length = 0
        segment_lines = 1
        for word in words:
            word_length = len(word)
            if line_length == 0:
                line_length = word_length
            elif line_length + 1 + word_length <= chars_per_line:
                line_length += 1 + word_length
            else:
                segment_lines += 1
                line_length = word_length

            if line_length > chars_per_line:
                extra_lines = (line_length - 1) // chars_per_line
                segment_lines += extra_lines
                line_length = line_length % chars_per_line or chars_per_line

        total_lines += segment_lines

    return max(1, total_lines)


def row_height_for_movement(value: object) -> int:
    line_count = estimate_wrapped_lines(value)
    if line_count <= 1:
        return BASE_ROW_HEIGHT
    return min(MAX_ROW_HEIGHT, ROW_HEIGHT_PADDING + WRAPPED_LINE_HEIGHT * line_count)


def export_columns(rows: list[dict[str, object]], treatment: Treatment) -> tuple[list[str], bool, list[int]]:
    include_refunds = any(row[treatment.secondary_column] not in (None, 0) for row in rows)
    headers = ["Libellé du compte", "Date", "Journal", "Libellé mouvement", treatment.primary_column]
    if include_refunds:
        headers.append(treatment.secondary_column)
    headers.append("Remarque")
    amount_columns = [headers.index(treatment.primary_column) + 1]
    if include_refunds:
        amount_columns.append(headers.index(treatment.secondary_column) + 1)
    return headers, include_refunds, amount_columns


def export_values(row: dict[str, object], treatment: Treatment, include_refunds: bool) -> list[object]:
    values = [
        row["Libellé du compte"],
        row["Date"],
        row["Journal"],
        row["Libellé mouvement"],
        row[treatment.primary_column],
    ]
    if include_refunds:
        values.append(row[treatment.secondary_column])
    values.append(row["Remarque"])
    return values


def read_rows(csv_path: Path, treatment: Treatment) -> list[dict[str, object]]:
    with csv_path.open("r", encoding="cp1252", newline="") as handle:
        reader = csv.DictReader(handle, delimiter=";")
        header_map = get_header_map(reader)

        rows: list[dict[str, object]] = []
        for raw in reader:
            piece = str(raw.get(header_map["n_de_piece"], "") or "").strip()
            journal = str(raw.get(header_map["journal"], "") or "").strip().upper()
            debit = parse_money(raw.get(header_map["montant_debit"]))
            credit = parse_money(raw.get(header_map["montant_credit"]))
            date_value = parse_date(raw.get(header_map["date"]))
            base = {
                "Libellé du compte": str(raw.get(header_map["libelle_du_compte"], "") or "").strip(),
                "Date": date_value,
                "Journal": journal,
                "Libellé mouvement": str(raw.get(header_map["libelle_mouvement"], "") or "").strip(),
            }

            if treatment.key == "supplier_payments":
                if piece or journal in EXCLUDED_JOURNALS:
                    continue
                rows.append(
                    {
                        **base,
                        "Paiements": debit if debit != 0 else None,
                        "Rmbrsmts": credit if credit != 0 else None,
                        "Remarque": "Manque facture" if debit > 0 else "Manque avoir",
                    }
                )
            elif treatment.key == "supplier_invoices":
                if journal not in {"AC", "AT"}:
                    continue
                rows.append(
                    {
                        **base,
                        "Factures": credit if credit != 0 else None,
                        "Avoirs": debit if debit != 0 else None,
                        "Remarque": "Facture sans paiement" if credit > 0 else "Avoir sans rmbrsmt",
                    }
                )
            elif treatment.key == "customer_receipts":
                if piece or journal in CUSTOMER_RECEIPT_EXCLUDED_JOURNALS:
                    continue
                rows.append(
                    {
                        **base,
                        "Encaissements": credit if credit != 0 else None,
                        "Remboursements": debit if debit != 0 else None,
                        "Remarque": "Manque facture de vente" if credit > 0 else "Manque avoir de vente",
                    }
                )
            elif treatment.key == "customer_invoices":
                if journal not in SALE_JOURNALS:
                    continue
                rows.append(
                    {
                        **base,
                        "Factures": debit if debit != 0 else None,
                        "Avoirs": credit if credit != 0 else None,
                        "Remarque": "Facture de vente sans paiement" if debit > 0 else "Avoir de vente sans rmbrsmt",
                    }
                )

    rows.sort(
        key=lambda row: (
            str(row["Libellé du compte"]).upper(),
            row["Date"] or datetime.max,
            str(row["Journal"]),
            float(row.get("Paiements") or row.get("Encaissements") or row.get("Factures") or 0),
            float(row.get("Rmbrsmts") or row.get("Remboursements") or row.get("Avoirs") or 0),
        )
    )
    return rows


def write_workbook(
    rows: list[dict[str, object]],
    output_path: Path,
    treatment: Treatment,
) -> None:
    headers, include_refunds, amount_columns = export_columns(rows, treatment)

    wb = Workbook()
    ws = wb.active
    ws.title = treatment.sheet_title
    last_col = len(headers)
    last_col_letter = get_column_letter(last_col)

    header_row = 1
    for column_index, header in enumerate(headers, start=1):
        ws.cell(header_row, column_index, header)

    for row_index, row in enumerate(rows, start=header_row + 1):
        values = export_values(row, treatment, include_refunds)
        for column_index, value in enumerate(values, start=1):
            ws.cell(row_index, column_index, value)

    last_row = max(header_row, header_row + len(rows))
    table_end_row = max(last_row, header_row + 1)
    table_ref = f"A{header_row}:{last_col_letter}{table_end_row}"
    table = Table(displayName=treatment.table_name, ref=table_ref)
    ws.add_table(table)

    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(name="Segoe UI", size=11, color="FFFFFF", bold=True)
    thin = Side(style="thin", color="808080")
    full_border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for cell in ws[header_row]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = full_border

    for row in ws.iter_rows(min_row=header_row + 1, max_row=last_row):
        for cell in row:
            horizontal = "left"
            if cell.column in (2, 3):
                horizontal = "center"
            elif cell.column in amount_columns:
                horizontal = "center"
            elif cell.column == last_col:
                horizontal = "center"
            cell.alignment = Alignment(horizontal=horizontal, vertical="center", wrap_text=cell.column == 4)
            cell.font = Font(name="Segoe UI", size=11)
            if cell.value not in (None, ""):
                cell.border = full_border

    widths = {
        "A": 24,
        "B": 12,
        "C": 9,
        "D": 76,
        "E": 16,
        "F": 16 if include_refunds else 18,
        "G": 18,
    }
    for col, width in widths.items():
        if last_col >= column_index_from_string(col):
            ws.column_dimensions[col].width = width

    for row_index in range(header_row + 1, last_row + 1):
        ws.cell(row_index, 2).number_format = "dd/mm/yyyy"
        for column_index in amount_columns:
            ws.cell(row_index, column_index).number_format = '#,##0.00 "€"'
        movement = str(ws.cell(row_index, 4).value or "")
        ws.row_dimensions[row_index].height = row_height_for_movement(movement)

    ws.freeze_panes = f"A{header_row + 1}"
    ws.sheet_view.showGridLines = False
    ws.page_setup.orientation = "portrait"
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.print_area = f"A{header_row}:{last_col_letter}{last_row}"
    ws.page_margins.left = 0.3
    ws.page_margins.right = 0.3
    ws.page_margins.top = 0.5
    ws.page_margins.bottom = 0.5
    ws.page_margins.footer = 0.2
    ws.oddFooter.center.text = "&P/&N"
    ws.evenFooter.center.text = "&P/&N"
    ws.firstFooter.center.text = "&P/&N"
    ws.sheet_view.zoomScale = 90

    wb.save(output_path)


def format_export_value(value: object, column_index: int, amount_columns: list[int]) -> str:
    if value in (None, ""):
        return ""
    if isinstance(value, datetime):
        return value.strftime("%d/%m/%Y")
    if column_index in amount_columns:
        return f"{format_amount(float(value))} €"
    return str(value)


def write_pdf(rows: list[dict[str, object]], output_path: Path, treatment: Treatment) -> None:
    try:
        from reportlab.lib import colors
        from reportlab.lib.enums import TA_CENTER, TA_LEFT
        from reportlab.lib.pagesizes import A4
        from reportlab.lib.styles import ParagraphStyle
        from reportlab.lib.units import mm
        from reportlab.pdfgen import canvas
        from reportlab.platypus import Paragraph, SimpleDocTemplate, Table as PdfTable, TableStyle
        from xml.sax.saxutils import escape
    except ImportError as exc:
        raise RuntimeError("La génération PDF nécessite reportlab.") from exc

    output_path.parent.mkdir(parents=True, exist_ok=True)
    headers, include_refunds, amount_columns = export_columns(rows, treatment)
    page_width, page_height = A4
    left_margin = right_margin = 8 * mm
    available_width = page_width - left_margin - right_margin

    if len(headers) == 7:
        weights = [1.2, 0.72, 0.48, 2.7, 0.92, 0.92, 1.2]
        font_size = 7.2
    else:
        weights = [1.35, 0.78, 0.52, 3.1, 1.05, 1.35]
        font_size = 7.6
    total_weight = sum(weights)
    col_widths = [available_width * weight / total_weight for weight in weights]

    header_style = ParagraphStyle(
        "Header",
        fontName="Helvetica-Bold",
        fontSize=font_size,
        leading=font_size + 1.5,
        alignment=TA_CENTER,
        textColor=colors.white,
    )
    left_style = ParagraphStyle(
        "BodyLeft",
        fontName="Helvetica",
        fontSize=font_size,
        leading=font_size + 1.6,
        alignment=TA_LEFT,
    )
    center_style = ParagraphStyle(
        "BodyCenter",
        parent=left_style,
        alignment=TA_CENTER,
    )

    data = [[Paragraph(escape(str(header)), header_style) for header in headers]]
    last_col = len(headers)
    for row in rows:
        values = export_values(row, treatment, include_refunds)
        pdf_row = []
        for column_index, value in enumerate(values, start=1):
            style = center_style if column_index in (2, 3, last_col) or column_index in amount_columns else left_style
            text = escape(format_export_value(value, column_index, amount_columns)).replace("\n", "<br/>")
            pdf_row.append(Paragraph(text, style))
        data.append(pdf_row)

    table = PdfTable(data, colWidths=col_widths, repeatRows=1)
    table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1F4E78")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("GRID", (0, 0), (-1, -1), 0.35, colors.HexColor("#808080")),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("LEFTPADDING", (0, 0), (-1, -1), 2),
                ("RIGHTPADDING", (0, 0), (-1, -1), 2),
                ("TOPPADDING", (0, 0), (-1, -1), 2),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 2),
            ]
        )
    )

    class NumberedCanvas(canvas.Canvas):
        def __init__(self, *args, **kwargs):
            super().__init__(*args, **kwargs)
            self._saved_page_states = []

        def showPage(self):
            self._saved_page_states.append(dict(self.__dict__))
            self._startPage()

        def save(self):
            page_count = len(self._saved_page_states)
            for state in self._saved_page_states:
                self.__dict__.update(state)
                self.setFont("Helvetica", 8)
                self.drawCentredString(page_width / 2, 7 * mm, f"{self._pageNumber}/{page_count}")
                super().showPage()
            super().save()

    doc = SimpleDocTemplate(
        str(output_path),
        pagesize=A4,
        leftMargin=left_margin,
        rightMargin=right_margin,
        topMargin=8 * mm,
        bottomMargin=14 * mm,
    )
    doc.build([table], canvasmaker=NumberedCanvas)


def process_file(
    csv_path: Path,
    treatment: Treatment,
    formats: OutputFormats = EXCEL_ONLY,
    root: Tk | None = None,
) -> OutputInfo:
    if not formats.has_any:
        raise ValueError("Sélectionnez au moins un format.")
    if csv_path.suffix.lower() != ".csv":
        raise ValueError(f"Le fichier n'est pas un CSV : {csv_path.name}")
    if not csv_path.exists():
        raise FileNotFoundError(str(csv_path))

    client, year = derive_client_year(csv_path, root)
    rows = read_rows(csv_path, treatment)
    total = sum(float(row.get(treatment.primary_column) or 0) for row in rows)
    if not rows:
        return OutputInfo(csv_path, None, None, 0, treatment.amount_label, total, treatment.label)

    output_stem = f"{client} - {treatment.output_label} {year}"
    excel_path, pdf_path = unique_output_paths(csv_path.parent, output_stem, formats)

    if formats.excel:
        write_workbook(rows, excel_path, treatment)
    if formats.pdf:
        write_pdf(rows, pdf_path, treatment)

    return OutputInfo(csv_path, excel_path, pdf_path, len(rows), treatment.amount_label, total, treatment.label)


def process_many(
    paths: list[Path],
    treatment_roles: tuple[str, ...],
    formats: OutputFormats = EXCEL_ONLY,
    root: Tk | None = None,
    ledger_mode: str = LEDGER_AUTO,
) -> list[OutputInfo]:
    if not paths:
        return []
    outputs = []
    errors = []
    for path in paths:
        if path.suffix.lower() != ".csv":
            errors.append(f"{path.name} : Le fichier n'est pas un CSV.")
            continue
        if not path.exists():
            errors.append(f"{path.name} : fichier introuvable.")
            continue
        try:
            treatments = resolve_treatments(path, treatment_roles, ledger_mode)
        except Exception as exc:
            errors.append(f"{path.name} : {exc}")
            continue
        for treatment in treatments:
            try:
                outputs.append(process_file(path, treatment, formats, root))
            except Exception as exc:
                errors.append(f"{path.name} ({treatment.label}) : {exc}")
    if errors:
        raise RuntimeError("\n".join(errors))
    return outputs


def success_text(outputs: list[OutputInfo]) -> str:
    created = [info for info in outputs if info.excel_output is not None or info.pdf_output is not None]
    skipped = [info for info in outputs if info.excel_output is None and info.pdf_output is None]
    created_file_count = sum(
        int(info.excel_output is not None) + int(info.pdf_output is not None)
        for info in created
    )
    lines = []
    if created:
        lines.append("Fichier généré :" if created_file_count == 1 else "Fichiers générés :")
        for info in created:
            lines.append(f"- {info.treatment_label} : {info.row_count} lignes, {info.amount_label} : {format_amount(info.total_amount)} €")
            if info.excel_output is not None:
                lines.append(f"  Excel : {info.excel_output}")
            if info.pdf_output is not None:
                lines.append(f"  PDF : {info.pdf_output}")
    if skipped:
        if lines:
            lines.append("")
        lines.append("Aucun fichier créé pour :")
        for info in skipped:
            lines.append(f"- {info.treatment_label} : aucune ligne trouvée")
    if not lines:
        lines.append("Aucun fichier généré.")
    return "\n".join(lines)


class App:
    def __init__(self) -> None:
        root_cls = TkinterDnD.Tk if TkinterDnD else Tk
        self.root = root_cls()
        self.root.title("Mise en forme - Grand livre")
        icon_path = resource_path("logo_compta_premium.ico")
        if icon_path.exists():
            try:
                self.root.iconbitmap(str(icon_path))
            except Exception:
                pass
        self.root.geometry("920x740")
        self.root.minsize(820, 700)
        self.root.configure(bg="#F5F7FA")
        saved_formats = load_saved_formats(EXCEL_ONLY, EXCEL_ONLY)
        saved_ledger_mode = load_saved_ledger_mode()
        self.excel_format_var = tk.BooleanVar(value=saved_formats.excel)
        self.pdf_format_var = tk.BooleanVar(value=saved_formats.pdf)
        self.ledger_mode_var = tk.StringVar(value=saved_ledger_mode)
        self.status_var = tk.StringVar(value="Déposez un CSV pour lancer un traitement.")
        self.drop_title_vars: list[tuple[tk.StringVar, tuple[str, ...]]] = []
        self.build_ui()

    def build_ui(self) -> None:
        title = tk.Label(
            self.root,
            text="Mise en forme du grand livre",
            font=("Segoe UI", 18, "bold"),
            bg="#F5F7FA",
            fg="#1F2937",
        )
        title.pack(pady=(22, 6))

        subtitle = tk.Label(
            self.root,
            text="Déposez un export CSV dans la zone correspondant au traitement souhaité.",
            font=("Segoe UI", 10),
            bg="#F5F7FA",
            fg="#4B5563",
        )
        subtitle.pack(pady=(0, 18))

        format_frame = tk.Frame(self.root, bg="#F5F7FA")
        format_frame.pack(pady=(0, 16))
        tk.Label(
            format_frame,
            text="Format de sortie",
            font=("Segoe UI", 10, "bold"),
            bg="#F5F7FA",
            fg="#334155",
        ).pack(side="left", padx=(0, 14))
        for text, variable in (("Excel", self.excel_format_var), ("PDF", self.pdf_format_var)):
            tk.Checkbutton(
                format_frame,
                text=text,
                variable=variable,
                command=self.save_selected_formats,
                font=("Segoe UI", 10),
                bg="#F5F7FA",
                fg="#1F2937",
                activebackground="#F5F7FA",
                activeforeground="#1F2937",
                selectcolor="#FFFFFF",
                cursor="hand2",
            ).pack(side="left", padx=8)

        mode_frame = tk.Frame(self.root, bg="#F5F7FA")
        mode_frame.pack(pady=(0, 16))
        tk.Label(
            mode_frame,
            text="Type de grand livre",
            font=("Segoe UI", 10, "bold"),
            bg="#F5F7FA",
            fg="#334155",
        ).pack(side="left", padx=(0, 14))
        for text, mode in (("Auto", LEDGER_AUTO), ("Fournisseur", LEDGER_SUPPLIER), ("Client", LEDGER_CUSTOMER)):
            tk.Radiobutton(
                mode_frame,
                text=text,
                value=mode,
                variable=self.ledger_mode_var,
                command=self.save_selected_ledger_mode,
                font=("Segoe UI", 10),
                bg="#F5F7FA",
                fg="#1F2937",
                activebackground="#F5F7FA",
                activeforeground="#1F2937",
                selectcolor="#FFFFFF",
                cursor="hand2",
            ).pack(side="left", padx=8)

        row_frame = tk.Frame(self.root, bg="#F5F7FA")
        row_frame.pack(fill="x", padx=34)
        row_frame.columnconfigure(0, weight=1)
        row_frame.columnconfigure(1, weight=1)

        self.create_drop_box(
            row_frame,
            treatment_roles=(ROLE_FIRST,),
            column=0,
            accent="#1F4E78",
        )
        self.create_drop_box(
            row_frame,
            treatment_roles=(ROLE_SECOND,),
            column=1,
            accent="#8A6A13",
        )

        bottom_frame = tk.Frame(self.root, bg="#F5F7FA")
        bottom_frame.pack(fill="x", padx=210, pady=(20, 0))
        bottom_frame.columnconfigure(0, weight=1)
        self.create_drop_box(
            bottom_frame,
            treatment_roles=(ROLE_FIRST, ROLE_SECOND),
            column=0,
            accent="#334155",
            compact=True,
        )

        note = "Le dépôt direct d'un CSV sur l'icône du .exe génère la première liste du type détecté."
        if not DND_FILES:
            note = "Le glisser-déposer dans la fenêtre est indisponible ; utilisez les boutons ou glissez sur le .exe."
        tk.Label(self.root, text=note, font=("Segoe UI", 9), bg="#F5F7FA", fg="#4B5563").pack(pady=(18, 0))

        self.status_frame = tk.Frame(
            self.root,
            bg="#FFFFFF",
            highlightbackground="#CBD5E1",
            highlightthickness=1,
            bd=0,
            height=105,
        )
        self.status_frame.pack(fill="x", padx=46, pady=(14, 0))
        self.status_frame.pack_propagate(False)

        self.status_title = tk.Label(
            self.status_frame,
            text="Résultat",
            font=("Segoe UI", 10, "bold"),
            bg="#FFFFFF",
            fg="#334155",
            anchor="w",
        )
        self.status_title.pack(fill="x", padx=14, pady=(10, 2))

        self.status_label = tk.Label(
            self.status_frame,
            textvariable=self.status_var,
            font=("Segoe UI", 9),
            bg="#FFFFFF",
            fg="#475569",
            justify="left",
            anchor="nw",
            wraplength=800,
        )
        self.status_label.pack(fill="both", expand=True, padx=14, pady=(0, 10))

    def title_for_roles(self, treatment_roles: tuple[str, ...]) -> str:
        if treatment_roles == (ROLE_FIRST, ROLE_SECOND):
            return "Générer les 2 fichiers"

        mode = self.ledger_mode_var.get()
        if treatment_roles == (ROLE_FIRST,):
            if mode == LEDGER_CUSTOMER:
                return "Lister les encaissements\nsans facture de ventes"
            if mode == LEDGER_SUPPLIER:
                return "Lister les paiements\nsans facture"
            return "Lister paiements / encaissements\nsans facture"

        if mode == LEDGER_CUSTOMER:
            return "Lister les factures de ventes\nsans paiement"
        if mode == LEDGER_SUPPLIER:
            return "Lister les factures\nsans paiements"
        return "Lister les factures\nsans paiements"

    def update_drop_titles(self) -> None:
        for title_var, treatment_roles in self.drop_title_vars:
            title_var.set(self.title_for_roles(treatment_roles))

    def create_drop_box(
        self,
        parent: tk.Frame,
        treatment_roles: tuple[str, ...],
        column: int,
        accent: str,
        compact: bool = False,
    ) -> None:
        frame = tk.Frame(
            parent,
            bg="#FFFFFF",
            highlightbackground=accent,
            highlightthickness=2,
            bd=0,
            height=125 if compact else 180,
            cursor="hand2",
        )
        frame.grid(row=0, column=column, padx=12, sticky="nsew")
        frame.pack_propagate(False)

        title_var = tk.StringVar(value=self.title_for_roles(treatment_roles))
        self.drop_title_vars.append((title_var, treatment_roles))
        label = tk.Label(
            frame,
            textvariable=title_var,
            font=("Segoe UI", 15 if compact else 16, "bold"),
            bg="#FFFFFF",
            fg=accent,
            justify="center",
            cursor="hand2",
        )
        label.pack(pady=(26 if compact else 42, 16))

        button = tk.Button(
            frame,
            text="Choisir un CSV",
            command=lambda: self.choose_file(treatment_roles),
            font=("Segoe UI", 10),
            bg=accent,
            fg="#FFFFFF",
            activebackground=accent,
            activeforeground="#FFFFFF",
            relief="flat",
            padx=14,
            pady=6,
            cursor="hand2",
        )
        button.pack()

        for widget in (frame, label):
            widget.bind("<Button-1>", lambda _event, selected=treatment_roles: self.choose_file(selected))

        if DND_FILES:
            for widget in (frame, label):
                widget.drop_target_register(DND_FILES)
                widget.dnd_bind("<<Drop>>", lambda event, selected=treatment_roles: self.on_drop(event, selected))

    def on_drop(self, event, treatment_roles: tuple[str, ...]) -> None:
        self.handle_paths([Path(item) for item in self.root.tk.splitlist(event.data)], treatment_roles)

    def choose_file(self, treatment_roles: tuple[str, ...]) -> None:
        files = filedialog.askopenfilenames(
            parent=self.root,
            title="Choisir un ou plusieurs CSV",
            filetypes=[("Fichiers CSV", "*.csv"), ("Tous les fichiers", "*.*")],
        )
        self.handle_paths([Path(file) for file in files], treatment_roles)

    def set_status(self, text: str, kind: str = "info") -> None:
        colors = {
            "info": ("#FFFFFF", "#CBD5E1", "#334155", "#475569"),
            "success": ("#F0FDF4", "#86EFAC", "#166534", "#14532D"),
            "error": ("#FEF2F2", "#FCA5A5", "#991B1B", "#7F1D1D"),
        }
        bg, border, title_fg, text_fg = colors.get(kind, colors["info"])
        self.status_frame.configure(bg=bg, highlightbackground=border)
        self.status_title.configure(bg=bg, fg=title_fg)
        self.status_label.configure(bg=bg, fg=text_fg)
        self.status_var.set(text)

    def selected_formats(self) -> OutputFormats:
        return OutputFormats(
            excel=bool(self.excel_format_var.get()),
            pdf=bool(self.pdf_format_var.get()),
        )

    def selected_ledger_mode(self) -> str:
        mode = str(self.ledger_mode_var.get() or LEDGER_AUTO)
        return mode if mode in LEDGER_MODES else LEDGER_AUTO

    def save_selected_formats(self) -> None:
        try:
            save_formats(self.selected_formats())
        except Exception as exc:
            self.set_status(f"Erreur lors de la sauvegarde du format :\n{exc}", "error")

    def save_selected_ledger_mode(self) -> None:
        try:
            save_ledger_mode(self.selected_ledger_mode())
            self.update_drop_titles()
        except Exception as exc:
            self.set_status(f"Erreur lors de la sauvegarde du type :\n{exc}", "error")

    def handle_paths(self, paths: list[Path], treatment_roles: tuple[str, ...]) -> None:
        if not paths:
            return
        formats = self.selected_formats()
        if not formats.has_any:
            self.set_status("Sélectionnez au moins un format.", "error")
            return
        try:
            outputs = process_many(paths, treatment_roles, formats, self.root, self.selected_ledger_mode())
            created = any(info.excel_output is not None or info.pdf_output is not None for info in outputs)
            self.set_status(success_text(outputs), "success" if created else "info")
        except Exception as exc:
            write_error_log(str(exc))
            self.set_status(f"Erreur :\n{exc}", "error")

    def run(self) -> None:
        self.root.mainloop()


def main() -> int:
    no_dialog = "--no-dialog" in sys.argv
    args = [Path(arg) for arg in sys.argv[1:] if arg != "--no-dialog"]
    if args:
        root = None if no_dialog else Tk()
        if root is not None:
            root.withdraw()
        try:
            formats = load_saved_formats(EXCEL_AND_PDF, EXCEL_ONLY)
            outputs = process_many(args, (ROLE_FIRST,), formats, root, LEDGER_AUTO)
            if not no_dialog:
                return 0
            if sys.stdout is not None:
                print(success_text(outputs))
            return 0
        except Exception as exc:
            write_error_log(str(exc))
            if no_dialog:
                if sys.stderr is not None:
                    print(str(exc), file=sys.stderr)
            else:
                messagebox.showerror("Erreur", str(exc), parent=root)
            return 1
        finally:
            if root is not None:
                root.destroy()

    App().run()
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
